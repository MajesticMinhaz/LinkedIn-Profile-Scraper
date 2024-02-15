from dotenv import dotenv_values
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from database import create_linkedin_data


# Get the config values from .env file
config = dotenv_values(dotenv_path='.env')


def read_excel_file(file_path: str, sheet_name: str):
    try:
        # Load the workbook
        workbook = load_workbook(file_path)

        # Select the desired sheet
        sheet = workbook[sheet_name]

        # Create a dictionary to store the data
        data = []

        # Get the column names (assuming the first row contains headers)
        columns = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            single_row_data = dict()
            for column_name, value in zip(columns, row):

                if str(value).strip() == '' or str(value).strip() == 'N/A' or str(value).strip() == 'None':
                    value = None

                    single_row_data[column_name] = value
                else:
                    single_row_data[column_name] = str(value).strip()

            data.append(single_row_data)

        return data

    except FileNotFoundError:
        print(f"File '{file_path}' not found.")
        return None
    except InvalidFileException:
        print(f"Invalid file format or corrupted file: '{file_path}'")
        return None


sheet_data = read_excel_file(
    file_path=config.get('INSERT_EXCEL_FILE_PATH'),
    sheet_name=config.get('INSERT_EXCEL_SHEET_NAME')
)


if sheet_data:
    for row_data in sheet_data:
        create_linkedin_data(
            profile_url=row_data.get('PROFILE URL'),
            first_name=row_data.get('FIRST NAME'),
            last_name=row_data.get('LAST NAME'),
            company_name=row_data.get('COMPANY NAME'),
            job_title=row_data.get('JOB TITLE')
        )

    print("Successfully Completed the task")

else:
    print("Reading the file failed.")
